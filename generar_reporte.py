import pandas as pd
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

print("Iniciando generación de reporte ejecutivo...")

# 1. Cargar la data limpia
try:
    df = pd.read_csv('datos_consolidados_limpios.csv')
    print(f"✅ Data cargada: {len(df)} registros")
except FileNotFoundError:
    print("❌ Error: No se encontró la data limpia. Ejecuta primero el pipeline de limpieza.")
    exit()

# 2. Validar columnas críticas
columnas_requeridas = ['estado_operativo', 'total_open_amount', 'invoice_currency', 'buisness_year']
columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
if columnas_faltantes:
    print(f"⚠️ Advertencia: Faltan columnas: {columnas_faltantes}")

# 3. Generar múltiples análisis
nombre_excel = 'reporte_ejecutivo_diario.xlsx'
with pd.ExcelWriter(nombre_excel, engine='openpyxl') as writer:
    
    # HOJA 1: KPIs Principales
    kpis: Dict[str, List[Any]] = {
        'KPI': [
            'Total de Documentos',
            'Documentos Pendientes',
            'Documentos Cerrados',
            '% Pendientes',
            '% Cerrados',
            'Monto Total Pendiente (USD)',
            'Monto Promedio por Documento'
        ],
        'Valor': [
            len(df),
            (df['estado_operativo'] == 'Pendiente').sum(),
            (df['estado_operativo'] == 'Cerrado').sum(),
            f"{((df['estado_operativo'] == 'Pendiente').sum() / len(df) * 100):.2f}%",
            f"{((df['estado_operativo'] == 'Cerrado').sum() / len(df) * 100):.2f}%",
            f"${df[df['estado_operativo'] == 'Pendiente']['total_open_amount'].sum():,.2f}",
            f"${df['total_open_amount'].mean():,.2f}"
        ]
    }
    df_kpis = pd.DataFrame(kpis)  # type: ignore[arg-type]
    df_kpis.to_excel(writer, sheet_name='KPIs Principales', index=False)  # type: ignore[union-attr]
    
    # HOJA 2: Resumen por Estado
    resumen_estado = df.groupby('estado_operativo').agg({
        'invoice_id': 'count',
        'total_open_amount': ['sum', 'mean', 'max']
    }).round(2)
    resumen_estado.columns = ['Cantidad', 'Monto Total', 'Monto Promedio', 'Monto Máximo']
    resumen_estado = resumen_estado.reset_index()
    resumen_estado.to_excel(writer, sheet_name='Resumen por Estado', index=False)  # type: ignore[union-attr]
    
    # HOJA 3: Análisis por Moneda
    if 'invoice_currency' in df.columns:
        resumen_moneda = df.groupby('invoice_currency').agg({
            'invoice_id': 'count',
            'total_open_amount': ['sum', 'mean']
        }).round(2)
        resumen_moneda.columns = ['Cantidad', 'Monto Total', 'Monto Promedio']
        resumen_moneda = resumen_moneda.reset_index()
        resumen_moneda.to_excel(writer, sheet_name='Por Moneda', index=False)  # type: ignore[union-attr]
    
    # HOJA 4: Análisis por Año
    if 'buisness_year' in df.columns:
        resumen_ano = df.groupby('buisness_year').agg({
            'invoice_id': 'count',
            'total_open_amount': 'sum'
        }).round(2)
        resumen_ano.columns = ['Cantidad', 'Monto Total']
        resumen_ano = resumen_ano.reset_index()
        resumen_ano.to_excel(writer, sheet_name='Por Año', index=False)  # type: ignore[union-attr]
    
    # HOJA 5: Top 10 Clientes
    if 'name_customer' in df.columns:
        top_clientes = df.groupby('name_customer').agg({
            'invoice_id': 'count',
            'total_open_amount': 'sum'
        }).sort_values('total_open_amount', ascending=False).head(10).round(2)
        top_clientes.columns = ['Cantidad Docs', 'Monto Total']
        top_clientes = top_clientes.reset_index()
        top_clientes.to_excel(writer, sheet_name='Top 10 Clientes', index=False)  # type: ignore[union-attr]
    
    # HOJA 6: Data Detallada (primeras 1000 filas)
    columnas_mostrar: List[str] = ['invoice_id', 'name_customer', 'estado_operativo', 'total_open_amount', 
                       'invoice_currency', 'posting_date', 'clear_date']
    columnas_validas = [col for col in columnas_mostrar if col in df.columns]
    df[columnas_validas].head(1000).to_excel(writer, sheet_name='Data Detallada', index=False)  # type: ignore[union-attr]

# 4. Formatear el Excel
wb = load_workbook(nombre_excel)

# Estilos
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    
    # Formatear encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter: str = column[0].column_letter  # type: ignore[assignment]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Aplicar bordes y alineación a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save(nombre_excel)

# 5. Mostrar resumen en consola
print("\n" + "="*60)
print("📊 RESUMEN DEL REPORTE EJECUTIVO")
print("="*60)
print(f"\n✅ Total de documentos: {len(df):,}")
print(f"📋 Pendientes: {(df['estado_operativo'] == 'Pendiente').sum():,} ({((df['estado_operativo'] == 'Pendiente').sum() / len(df) * 100):.1f}%)")
print(f"✔️  Cerrados: {(df['estado_operativo'] == 'Cerrado').sum():,} ({((df['estado_operativo'] == 'Cerrado').sum() / len(df) * 100):.1f}%)")
print(f"💰 Monto pendiente: ${df[df['estado_operativo'] == 'Pendiente']['total_open_amount'].sum():,.2f}")
print(f"📅 Fecha del reporte: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"\n📁 Reporte guardado como: {nombre_excel}")
print("="*60 + "\n")